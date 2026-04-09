from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Test Cases"

def side(): return Side(style="thin", color="CBD5E1")
def border(): return Border(left=side(), right=side(), top=side(), bottom=side())
def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def wrap_align(h="left"): return Alignment(wrap_text=True, vertical="top", horizontal=h)

COLUMNS = [
    ("TC ID", 12), ("Module", 20), ("Test Scenario", 38),
    ("Test Steps", 55), ("Expected Result", 50),
    ("Actual Result", 28), ("Status", 12), ("Priority", 12),
]
for i, (_, w) in enumerate(COLUMNS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title
ws.merge_cells("A1:H1")
t = ws["A1"]
t.value = "VABERP — Complete QA Test Cases  |  185 Test Cases  |  v2.0  |  2026-04-07"
t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t.fill = fill("1E293B")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Header row
for col, (name, _) in enumerate(COLUMNS, 1):
    c = ws.cell(row=2, column=col, value=name)
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    c.fill = fill("334155")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border()
ws.row_dimensions[2].height = 22

SECTION_LABELS = {
    "A": "SECTION A — Authentication — All Roles (TC-001 to TC-018)",
    "B": "SECTION B — Super Admin Module (TC-019 to TC-040)",
    "C": "SECTION C — Reseller Admin Module (TC-041 to TC-058)",
    "D": "SECTION D — Role-Based Access Control / RBAC (TC-059 to TC-067)",
    "E": "SECTION E — Manager Hierarchy & Visibility (TC-068 to TC-073)",
    "F": "SECTION F — RFQ Module (TC-074 to TC-086)",
    "G": "SECTION G — Quotation Module (TC-087 to TC-110)",
    "H": "SECTION H — Purchase Order Module (TC-111 to TC-125)",
    "I": "SECTION I — Invoice Module (TC-126 to TC-139)",
    "J": "SECTION J — Sales Module (TC-140 to TC-147)",
    "K": "SECTION K — Employee & Role Management (TC-148 to TC-157)",
    "L": "SECTION L — UI, Navigation & Forms (TC-158 to TC-169)",
    "M": "SECTION M — API Testing (TC-170 to TC-180)",
    "N": "SECTION N — Negative & Edge Cases (TC-181 to TC-185)",
}

PRIORITY_FILL = {"High": "FEE2E2", "Medium": "FEF9C3", "Low": "DCFCE7"}

test_cases = [
    # ── SECTION A — Authentication All Roles ──────────────────────────────────
    ("A","TC-001","Auth","Super Admin login with valid credentials","1. Go to /superadmin/login\n2. Enter super admin email & password\n3. Click Login","Super Admin dashboard loads with full platform menu (Enterprises, Resellers, Subscriptions, Support, etc.)","","","High"),
    ("A","TC-002","Auth","Reseller Admin login with valid credentials","1. Go to /reseller/login\n2. Enter reseller email & password\n3. Click Login","Reseller dashboard loads with tenant management menu (Tenants, Plans, Wallet, Commissions, etc.)","","","High"),
    ("A","TC-003","Auth","Enterprise Admin login with valid credentials","1. Go to /login\n2. Enter admin email & password\n3. Click Login","Admin dashboard loads with full enterprise menu","","","High"),
    ("A","TC-004","Auth","Manager login with valid credentials","1. Go to /login\n2. Enter manager credentials\n3. Click Login","Manager dashboard loads; team-scoped menu visible","","","High"),
    ("A","TC-005","Auth","Employee login with valid credentials","1. Go to /login\n2. Enter employee credentials\n3. Click Login","Employee dashboard loads; restricted menu (no PO create, no approvals)","","","High"),
    ("A","TC-006","Auth","Super Admin login with wrong password","1. Go to /superadmin/login\n2. Enter wrong password\n3. Click Login","Error: Invalid credentials. No redirect. Token not issued","","","High"),
    ("A","TC-007","Auth","Reseller login with wrong password","1. Go to /reseller/login\n2. Enter wrong password\n3. Click Login","Error: Invalid credentials. No access granted","","","High"),
    ("A","TC-008","Auth","Super Admin token cannot access Enterprise routes","1. Login as Super Admin\n2. Call GET /quotations (enterprise API) with super admin JWT","401 or 403 — wrong token type. Super admin JWT is not valid on enterprise routes","","","High"),
    ("A","TC-009","Auth","Reseller token cannot access Super Admin routes","1. Login as Reseller\n2. Call GET /super-admin/enterprises with reseller JWT","403 Forbidden — ResellerGuard does not pass SuperAdminGuard","","","High"),
    ("A","TC-010","Auth","Employee token cannot access Reseller routes","1. Login as Employee\n2. Call GET /resellers/me/tenants with employee JWT","403 Forbidden — wrong token type","","","High"),
    ("A","TC-011","Auth","Login with blank fields — all portals","1. Open any login page\n2. Leave email and password blank\n3. Click Login","Frontend validation errors shown on both fields. No API call made","","","Medium"),
    ("A","TC-012","Auth","Session persistence after refresh — Super Admin","1. Login as Super Admin\n2. Refresh browser","Super Admin remains logged in. Dashboard reloads correctly","","","High"),
    ("A","TC-013","Auth","Session persistence after refresh — Reseller","1. Login as Reseller\n2. Refresh browser","Reseller remains logged in. Dashboard reloads correctly","","","High"),
    ("A","TC-014","Auth","Logout clears session — Super Admin","1. Login as Super Admin\n2. Click Logout","Session cleared. Redirected to /superadmin/login. Back button does not return to dashboard","","","High"),
    ("A","TC-015","Auth","Logout clears session — Reseller","1. Login as Reseller\n2. Click Logout","Session cleared. Redirected to /reseller/login","","","High"),
    ("A","TC-016","Auth","Locked Super Admin account cannot login","1. Super Admin locks a reseller account\n2. That reseller tries to login","Login rejected: Account is locked or inactive","","","High"),
    ("A","TC-017","Auth","SQL injection in Super Admin login","1. Go to /superadmin/login\n2. Enter ' OR 1=1 -- in email field\n3. Click Login","Login rejected. No data exposed. Standard error shown","","","High"),
    ("A","TC-018","Auth","Expired reseller subscription — login still works but features restricted","1. Reseller subscription is expired\n2. Login as that Reseller","Login succeeds but platform shows subscription expired warning. Tenant creation blocked","","","Medium"),

    # ── SECTION B — Super Admin Module ───────────────────────────────────────
    ("B","TC-019","Super Admin","Super Admin dashboard loads platform KPIs","1. Login as Super Admin\n2. Open /superadmin/dashboard","Dashboard shows: Total Enterprises, Active Enterprises, Total Resellers, Platform Revenue, Monthly Charts","","","High"),
    ("B","TC-020","Super Admin","Super Admin views all enterprises","1. Login as Super Admin\n2. Navigate to /superadmin/enterprises","All enterprises listed with name, status, subscription expiry, creation date","","","High"),
    ("B","TC-021","Super Admin","Super Admin creates a new enterprise","1. Login as Super Admin\n2. Navigate to Enterprises\n3. Click Create\n4. Fill details and assign to Reseller\n5. Save","New enterprise created. Appears in enterprises list. Assigned reseller can see it in their tenants","","","High"),
    ("B","TC-022","Super Admin","Super Admin views enterprise financial details","1. Login as Super Admin\n2. Open an enterprise\n3. Change period to 90 days","Revenue, costs, profit margins, invoice/PO/SO counts all shown for 90-day period","","","High"),
    ("B","TC-023","Super Admin","Super Admin approves an enterprise","1. Login as Super Admin\n2. Open a pending enterprise\n3. Click Approve","Enterprise status changes to Active. Enterprise admin can now login","","","High"),
    ("B","TC-024","Super Admin","Super Admin locks an enterprise","1. Login as Super Admin\n2. Open an active enterprise\n3. Click Lock","Enterprise status = Locked. Enterprise admin login is blocked","","","High"),
    ("B","TC-025","Super Admin","Super Admin unlocks an enterprise","1. Login as Super Admin\n2. Open a locked enterprise\n3. Click Unlock","Enterprise status = Active. Enterprise admin login is restored","","","High"),
    ("B","TC-026","Super Admin","Super Admin reassigns enterprise to different reseller","1. Login as Super Admin\n2. Open an enterprise\n3. Change reseller assignment\n4. Save","Enterprise now appears under new reseller's tenant list. Previous reseller no longer sees it","","","High"),
    ("B","TC-027","Super Admin","Super Admin manages resellers list","1. Login as Super Admin\n2. Navigate to /superadmin/resellers","All resellers listed with name, status, plan, wallet balance, tenant count","","","High"),
    ("B","TC-028","Super Admin","Super Admin creates a new reseller","1. Login as Super Admin\n2. Navigate to Resellers\n3. Click Create\n4. Fill reseller details\n5. Save","New reseller account created. Reseller can login at /reseller/login","","","High"),
    ("B","TC-029","Super Admin","Super Admin locks a reseller account","1. Login as Super Admin\n2. Open a reseller\n3. Click Lock","Reseller status = Locked. Reseller cannot login. Their tenants are unaffected","","","High"),
    ("B","TC-030","Super Admin","Super Admin credits reseller wallet","1. Login as Super Admin\n2. Open Resellers > Wallets\n3. Select reseller\n4. Credit 5000 to wallet","Reseller wallet balance increases by 5000. Transaction record created","","","High"),
    ("B","TC-031","Super Admin","Super Admin creates a subscription plan","1. Login as Super Admin\n2. Navigate to Subscriptions\n3. Create new plan with price, features, limits\n4. Save","Plan created and available for resellers to assign to tenants","","","High"),
    ("B","TC-032","Super Admin","Super Admin manages reseller plans","1. Login as Super Admin\n2. Navigate to Resellers > Plans\n3. Create plan with commission % and max tenants","Plan created. Resellers can subscribe to this plan","","","High"),
    ("B","TC-033","Super Admin","Super Admin views all support tickets","1. Login as Super Admin\n2. Navigate to /superadmin/support","All support tickets from all enterprises and resellers listed","","","High"),
    ("B","TC-034","Super Admin","Super Admin replies to a support ticket","1. Login as Super Admin\n2. Open a ticket\n3. Type reply\n4. Submit","Reply saved. Ticket status updated. Requester notified","","","High"),
    ("B","TC-035","Super Admin","Super Admin manages services catalog","1. Login as Super Admin\n2. Navigate to /superadmin/services\n3. Add/edit a service","Service catalog updated. Reflects across all enterprise plans","","","Medium"),
    ("B","TC-036","Super Admin","Super Admin creates a coupon code","1. Login as Super Admin\n2. Navigate to /superadmin/coupons\n3. Create coupon with discount % and expiry\n4. Save","Coupon created. Can be applied during enterprise subscription","","","Medium"),
    ("B","TC-037","Super Admin","Super Admin views platform-level accounts","1. Login as Super Admin\n2. Navigate to /superadmin/accounts\n3. Change period filter to 365 days","Platform-level revenue summary shown for selected period","","","Medium"),
    ("B","TC-038","Super Admin","Super Admin views all employees across platform","1. Login as Super Admin\n2. Navigate to /superadmin/employees","All employees across all enterprises listed with name, role, and enterprise","","","Medium"),
    ("B","TC-039","Super Admin","Super Admin updates enterprise subscription expiry","1. Login as Super Admin\n2. Open an enterprise\n3. Update subscription expiry date","Expiry date saved. Enterprise subscription extended","","","High"),
    ("B","TC-040","Super Admin","Super Admin cannot access Enterprise/Employee ERP modules","1. Login as Super Admin\n2. Try to navigate to /quotations or /purchase-orders","Access denied. Super Admin portal is completely separate from ERP modules","","","High"),

    # ── SECTION C — Reseller Admin Module ────────────────────────────────────
    ("C","TC-041","Reseller Admin","Reseller dashboard loads KPIs","1. Login as Reseller\n2. Open /reseller/dashboard","Shows: Total Tenants, Active Subscriptions, Expired Subscriptions, Revenue, Commission Earned, Wallet Balance","","","High"),
    ("C","TC-042","Reseller Admin","Reseller views all their tenants","1. Login as Reseller\n2. Navigate to /reseller/tenants","Only enterprises assigned to this reseller are listed","","","High"),
    ("C","TC-043","Reseller Admin","Reseller creates a new tenant","1. Login as Reseller\n2. Navigate to Tenants\n3. Click Create\n4. Fill enterprise/tenant details\n5. Assign a plan\n6. Save","New tenant created under this reseller. Tenant admin can now login","","","High"),
    ("C","TC-044","Reseller Admin","Reseller cannot see other resellers tenants","1. Login as Reseller A\n2. Open /reseller/tenants","Only Reseller A's tenants visible. Reseller B's tenants NOT shown","","","High"),
    ("C","TC-045","Reseller Admin","Reseller assigns subscription plan to tenant","1. Login as Reseller\n2. Open a tenant\n3. Assign a subscription plan\n4. Save","Plan assigned to tenant. Tenant subscription activated with expiry date","","","High"),
    ("C","TC-046","Reseller Admin","Reseller renews expiring tenant subscription","1. Login as Reseller\n2. Open Subscriptions list\n3. Find expiring tenant\n4. Click Renew","Subscription extended. New expiry date set. Wallet debited or invoice created","","","High"),
    ("C","TC-047","Reseller Admin","Reseller views wallet balance and transactions","1. Login as Reseller\n2. Navigate to /reseller/wallet","Wallet balance shown. Transaction history (credits, debits, commissions) listed","","","High"),
    ("C","TC-048","Reseller Admin","Reseller views commission earnings","1. Login as Reseller\n2. Navigate to /reseller/commissions","Commission earnings listed by tenant/plan/date. Total commission shown","","","High"),
    ("C","TC-049","Reseller Admin","Reseller sets custom pricing for a plan","1. Login as Reseller\n2. Navigate to /reseller/plans\n3. Select a plan\n4. Set custom price\n5. Save","Custom price saved. Tenants of this reseller are charged the custom price","","","High"),
    ("C","TC-050","Reseller Admin","Reseller views usage per tenant","1. Login as Reseller\n2. Navigate to /reseller/usage","Employee count, active vs total employees shown per tenant","","","Medium"),
    ("C","TC-051","Reseller Admin","Reseller views subscription status of all tenants","1. Login as Reseller\n2. Navigate to /reseller/subscriptions","All tenant subscriptions listed with status (Active, Expired, Expiring Soon)","","","High"),
    ("C","TC-052","Reseller Admin","Reseller views their own subscription plan","1. Login as Reseller\n2. Navigate to /reseller/my-subscription","Own reseller plan details shown: plan name, max tenants, commission %, expiry","","","Medium"),
    ("C","TC-053","Reseller Admin","Reseller views billing/invoice history","1. Login as Reseller\n2. Navigate to /reseller/billing","All billing invoices listed with amounts and payment status","","","Medium"),
    ("C","TC-054","Reseller Admin","Reseller views reports","1. Login as Reseller\n2. Navigate to /reseller/reports","Sales/revenue/commission/tenant activity reports shown for own portfolio","","","Medium"),
    ("C","TC-055","Reseller Admin","Reseller updates own profile","1. Login as Reseller\n2. Navigate to /reseller/profile\n3. Update name and company\n4. Save","Profile updated. Changes saved and reflected in header/account area","","","Low"),
    ("C","TC-056","Reseller Admin","Reseller cannot access Super Admin portal","1. Login as Reseller\n2. Navigate to /superadmin/dashboard","Access denied or redirected to /reseller/login","","","High"),
    ("C","TC-057","Reseller Admin","Reseller cannot access Enterprise ERP modules","1. Login as Reseller\n2. Navigate to /quotations or /purchase-orders","Access denied. Reseller portal is completely separate from ERP modules","","","High"),
    ("C","TC-058","Reseller Admin","Reseller max tenant limit enforced","1. Reseller plan has max 5 tenants\n2. Reseller already has 5 tenants\n3. Try to create 6th tenant","Error: Maximum tenant limit reached for your plan. Upgrade plan to add more","","","High"),

    # ── SECTION D — RBAC ─────────────────────────────────────────────────────
    ("D","TC-059","RBAC","Employee cannot access PO creation","1. Login as Employee\n2. Navigate to /purchase-orders/new","Access denied or button not visible. No PO form rendered","","","High"),
    ("D","TC-060","RBAC","Employee cannot generate invoices","1. Login as Employee\n2. Open a Received PO\n3. Check for Generate Invoice action","Generate Invoice button not visible to Employee","","","High"),
    ("D","TC-061","RBAC","Employee cannot access User Management","1. Login as Employee\n2. Navigate to /settings/users via URL","Redirect to dashboard or 403 error. No user list shown","","","High"),
    ("D","TC-062","RBAC","Manager cannot access User Management","1. Login as Manager\n2. Navigate to /settings/users","Access denied. Create/Edit user actions unavailable","","","High"),
    ("D","TC-063","RBAC","Manager can approve quotations","1. Login as Manager\n2. Open Submitted quotation\n3. Click Approve","Approval action succeeds. Status changes to Approved","","","High"),
    ("D","TC-064","RBAC","Employee cannot approve quotations","1. Login as Employee\n2. Open a Submitted quotation","Approve button not visible. Cannot trigger approval action","","","High"),
    ("D","TC-065","RBAC","Admin can access all ERP modules","1. Login as Enterprise Admin\n2. Navigate through every menu item","All ERP modules accessible without restriction","","","High"),
    ("D","TC-066","RBAC","Direct URL access by unauthorized role","1. Login as Employee\n2. Manually type /purchase-orders/new in URL bar","Redirected away or 403 error shown. No form rendered","","","High"),
    ("D","TC-067","RBAC","API call with Employee token to Manager-only endpoint","1. Login as Employee via Postman\n2. Call POST /purchase-orders with employee JWT","403 Forbidden response returned","","","High"),

    # ── SECTION E — Manager Hierarchy ─────────────────────────────────────────
    ("E","TC-068","Hierarchy","Manager A sees Employee X and Y records","1. Employee X and Y (both under Manager A) create quotations\n2. Login as Manager A\n3. Open Quotations list","Records of both Employee X and Employee Y visible to Manager A","","","High"),
    ("E","TC-069","Hierarchy","Manager A cannot see Employee Z records (under B)","1. Employee Z (under Manager B) creates a quotation\n2. Login as Manager A","Employee Z quotation NOT in Manager A list","","","High"),
    ("E","TC-070","Hierarchy","Manager B cannot see Manager A team records","1. Employee X (under Manager A) creates a record\n2. Login as Manager B","Employee X record not visible to Manager B","","","High"),
    ("E","TC-071","Hierarchy","Admin sees all records from all users","1. Employees under Manager A and B each create records\n2. Login as Admin","All records from all users visible","","","High"),
    ("E","TC-072","Hierarchy","Employee reassigned — visibility updates immediately","1. Employee X under Manager A\n2. Admin reassigns to Manager B\n3. Login as Manager B","Employee X records visible to Manager B. No longer visible to Manager A","","","High"),
    ("E","TC-073","Hierarchy","Manager own records visible to themselves","1. Manager A creates a quotation directly\n2. Login as Manager A","Manager A own quotation visible in their list","","","Medium"),

    # ── SECTION F — RFQ ──────────────────────────────────────────────────────
    ("F","TC-074","RFQ","Create RFQ with valid data","1. Login\n2. Navigate to RFQ\n3. Fill vendor, items, quantities\n4. Submit","RFQ created with status Draft/Sent and unique reference number","","","High"),
    ("F","TC-075","RFQ","Create RFQ with missing vendor","1. Fill RFQ items\n2. Leave vendor blank\n3. Submit","Validation error: Vendor is required","","","High"),
    ("F","TC-076","RFQ","Create RFQ with zero quantity","1. Create RFQ\n2. Set item qty = 0\n3. Submit","Validation error: Quantity must be greater than 0","","","Medium"),
    ("F","TC-077","RFQ","Create RFQ with negative quantity","1. Create RFQ\n2. Enter qty = -5\n3. Submit","Validation error: negative values not allowed","","","Medium"),
    ("F","TC-078","RFQ","Create RFQ with no line items","1. Fill vendor\n2. Add no items\n3. Submit","Validation error: At least one line item is required","","","High"),
    ("F","TC-079","RFQ","Convert RFQ to Quotation","1. Open a submitted/received RFQ\n2. Click Convert to Quotation","Quotation created at v1 with all items pre-filled from RFQ","","","High"),
    ("F","TC-080","RFQ","Quotation from RFQ pre-fills all line items","1. Create RFQ with 3 items\n2. Convert to Quotation\n3. Open the quotation","All 3 items present in quotation with matching quantity and description","","","High"),
    ("F","TC-081","RFQ","Employee sees only their own RFQs","1. Employee X creates an RFQ\n2. Login as Employee Y","Employee Y cannot see Employee X RFQ","","","High"),
    ("F","TC-082","RFQ","Manager sees all team RFQs","1. Employee X (under Manager A) creates an RFQ\n2. Login as Manager A","RFQ appears in Manager A list","","","High"),
    ("F","TC-083","RFQ","Duplicate RFQ same vendor same reference","1. Create RFQ for Vendor A ref #001\n2. Create another with same vendor and reference","System warns or blocks exact duplicate","","","Medium"),
    ("F","TC-084","RFQ","RFQ list displays correct status badges","1. Create RFQs in different states\n2. Open RFQ list","Status badges correctly show Draft, Sent, Received per record","","","Low"),
    ("F","TC-085","RFQ","Delete a Draft RFQ","1. Create an RFQ\n2. Keep it in Draft\n3. Delete it","RFQ removed from list. No linked records affected","","","Medium"),
    ("F","TC-086","RFQ","Cannot delete a converted RFQ","1. Convert RFQ to Quotation\n2. Try to delete the original RFQ","Deletion blocked: RFQ has a linked Quotation","","","Medium"),

    # ── SECTION G — Quotation ─────────────────────────────────────────────────
    ("G","TC-087","Quotation","Create quotation with valid line items","1. Navigate to Create Quotation\n2. Add items with qty, unit price, discount\n3. Save","Quotation saved at v1, status = Draft, total calculated correctly","","","High"),
    ("G","TC-088","Quotation","Total calculation — single item with discount","1. Add Item: qty=2, price=1000, discount=10%","Total = (2x1000) - 10% = 1800","","","High"),
    ("G","TC-089","Quotation","Total calculation — multiple items","1. Item A: qty=1, price=500, 0% discount\n2. Item B: qty=2, price=300, 5% discount","Total = 500 + 570 = 1070","","","High"),
    ("G","TC-090","Quotation","Discount over 100%","1. Create quotation\n2. Enter discount = 101%\n3. Save","Validation error: Discount cannot exceed 100%","","","High"),
    ("G","TC-091","Quotation","Create quotation with no line items","1. Open Create Quotation\n2. Add no line items\n3. Submit","Validation error: At least one line item is required","","","High"),
    ("G","TC-092","Quotation","Submit quotation for approval","1. Open a Draft quotation\n2. Click Submit","Status = Submitted. Approve/Reject options now available to approvers","","","High"),
    ("G","TC-093","Quotation","Approve a submitted quotation","1. Login as Manager or Admin\n2. Open Submitted quotation\n3. Click Approve","Status = Approved. Version number unchanged. Edit fields locked","","","High"),
    ("G","TC-094","Quotation","Reject a submitted quotation","1. Login as Manager or Admin\n2. Open Submitted quotation\n3. Click Reject\n4. Enter rejection reason","Status = Rejected. Version number stays the same (v1 stays v1)","","","High"),
    ("G","TC-095","Quotation","CRITICAL: Version does NOT increment on rejection","1. Create Quotation (version = 1)\n2. Submit\n3. Manager rejects\n4. Check version field","Version field = 1 unchanged. Rejection must NEVER auto-increment version","","","High"),
    ("G","TC-096","Quotation","Revise a rejected quotation","1. Open Rejected quotation\n2. Click Revise\n3. Modify at least one line item\n4. Resubmit","New version created: v1 to v2. Original v1 preserved in version history","","","High"),
    ("G","TC-097","Quotation","CRITICAL: Version increments on user revision","1. Reject a v1 quotation\n2. User clicks Revise\n3. Check version after save","Version = 2. Re-submission shows v2 to approver","","","High"),
    ("G","TC-098","Quotation","Multiple revision cycles — correct versioning","1. v1 Reject Revise v2 Reject Revise v3","Versions: v1, v2, v3 in correct sequence. All listed in history","","","High"),
    ("G","TC-099","Quotation","Approved quotation fields are read-only","1. Open an Approved quotation\n2. Try to click and edit any field","All fields read-only. Only Revise action button available","","","High"),
    ("G","TC-100","Quotation","Revise approved quotation creates new version","1. Open Approved quotation\n2. Click Revise\n3. Edit and resubmit","New version (e.g. v2) created in Draft. Previous approved version preserved","","","High"),
    ("G","TC-101","Quotation","Employee cannot approve own submitted quotation","1. Login as Employee\n2. Create and submit a quotation\n3. Check for Approve button","Approve button NOT visible to the Employee who submitted it","","","High"),
    ("G","TC-102","Quotation","Quotation linked to correct customer","1. Create quotation for Customer A\n2. Open quotation detail","Customer A shown in customer field. No data leakage from other customers","","","High"),
    ("G","TC-103","Quotation","Search quotation by customer name","1. Open Quotation list\n2. Search by customer name","Only quotations linked to that customer are shown","","","Medium"),
    ("G","TC-104","Quotation","Filter quotations by status","1. Open Quotation list\n2. Apply filter Approved","Only Approved quotations displayed in list","","","Medium"),
    ("G","TC-105","Quotation","Cannot re-submit already Submitted quotation","1. Submit a quotation (status = Submitted)\n2. Click Submit again","Submit button disabled or already-submitted message shown","","","Medium"),
    ("G","TC-106","Quotation","Quotation PDF/print generation","1. Open an Approved quotation\n2. Click Download or Print","PDF generated with correct line items, totals, customer details, and version","","","Medium"),
    ("G","TC-107","Quotation","Duplicate quotation reference prevention","1. Create Quotation ref #Q-001 for Customer A\n2. Attempt to create another with same ref/customer","System warns or prevents exact duplicate quotation creation","","","Medium"),
    ("G","TC-108","Quotation","Version history shows all revisions with metadata","1. Create quotation with 3 revision cycles\n2. Open Version History tab","v1, v2, v3 listed with timestamp, status, and who approved/rejected","","","Medium"),
    ("G","TC-109","Quotation","Discount exactly 100%","1. Create quotation\n2. Enter discount = 100% on a line item","Accepted — line item cost = 0. Total reflects correctly","","","Medium"),
    ("G","TC-110","Quotation","Discount = 0% no discount","1. Leave discount blank or enter 0","No discount applied. Full price used in total","","","Low"),

    # ── SECTION H — Purchase Order ────────────────────────────────────────────
    ("H","TC-111","Purchase Order","Create PO from approved quotation","1. Open Approved quotation\n2. Click Create PO\n3. Assign to Employee\n4. Save","PO created with correct amounts and assigned to selected employee","","","High"),
    ("H","TC-112","Purchase Order","PO amount matches approved quotation total","1. Approved quotation total = 25000\n2. Create PO from it","PO total = 25000. No discrepancy between quotation and PO","","","High"),
    ("H","TC-113","Purchase Order","Cannot create PO from Draft quotation","1. Open a Draft quotation\n2. Check for Create PO option","Create PO button not visible or disabled","","","High"),
    ("H","TC-114","Purchase Order","Cannot create PO from Rejected quotation","1. Open a Rejected quotation\n2. Check for Create PO option","Create PO button not available for rejected quotations","","","High"),
    ("H","TC-115","Purchase Order","Cannot create PO from Submitted quotation","1. Open a Submitted quotation\n2. Check Create PO","Create PO disabled — must wait for approval","","","High"),
    ("H","TC-116","Purchase Order","Duplicate PO prevention for same quotation","1. Create PO from Approved quotation\n2. Try to create another PO from same quotation","System blocks or warns: A PO already exists for this quotation","","","High"),
    ("H","TC-117","Purchase Order","Employee sees only their assigned POs","1. Create PO assigned to Employee X\n2. Login as Employee X\n3. Open PO list","PO assigned to Employee X is visible","","","High"),
    ("H","TC-118","Purchase Order","Employee cannot see PO assigned to another employee","1. Create PO assigned to Employee Y\n2. Login as Employee X\n3. Open PO list","Employee Y PO is NOT in Employee X list","","","High"),
    ("H","TC-119","Purchase Order","Manager sees all team POs","1. Create POs for Employee X and Y (both under Manager A)\n2. Login as Manager A","Both POs from Employee X and Y visible to Manager A","","","High"),
    ("H","TC-120","Purchase Order","Update PO status to Received","1. Open an active PO\n2. Update status to Received","Status updates to Received. Invoice generation now enabled","","","High"),
    ("H","TC-121","Purchase Order","PO cannot be deleted if linked invoice exists","1. Generate an invoice from PO\n2. Try to delete the PO","Deletion blocked: Cannot delete a PO with an associated invoice","","","High"),
    ("H","TC-122","Purchase Order","Search PO by vendor name","1. Open PO list\n2. Search Vendor A","Only POs linked to Vendor A shown","","","Medium"),
    ("H","TC-123","Purchase Order","PO reference number is unique","1. Create multiple POs","Each PO has a unique system-generated reference number","","","High"),
    ("H","TC-124","Purchase Order","Reassign PO updates visibility","1. PO assigned to Employee X\n2. Edit PO and reassign to Employee Y","PO now visible to Employee Y. No longer visible to Employee X","","","Medium"),
    ("H","TC-125","Purchase Order","PO list sorted by date newest first","1. Create 3 POs on different dates\n2. Open PO list","Most recently created PO appears first","","","Low"),

    # ── SECTION I — Invoice ───────────────────────────────────────────────────
    ("I","TC-126","Invoice","Generate invoice from Received PO","1. Open PO with status = Received\n2. Click Generate Invoice","Invoice created with amount pre-filled from PO. Status = Unpaid","","","High"),
    ("I","TC-127","Invoice","Invoice amount auto-populated from PO","1. PO total = 40000\n2. Generate invoice","Invoice amount = 40000. No manual entry required","","","High"),
    ("I","TC-128","Invoice","Invoice default status is Unpaid","1. Generate any new invoice","Invoice status = Unpaid on creation","","","High"),
    ("I","TC-129","Invoice","Record partial payment","1. Invoice total = 50000\n2. Record payment of 20000","Status = Partially Paid. Remaining balance = 30000 shown","","","High"),
    ("I","TC-130","Invoice","Record full payment","1. Invoice total = 50000\n2. Record payment of 50000","Status = Paid. Balance = 0. Invoice marked complete","","","High"),
    ("I","TC-131","Invoice","Payment amount exceeds invoice total","1. Invoice = 50000\n2. Attempt to record 60000 payment","Validation error: Payment cannot exceed invoice total","","","High"),
    ("I","TC-132","Invoice","Multiple partial payments until fully paid","1. Invoice = 50000\n2. Record 20000 then 20000 then 10000","Each payment recorded. Final payment triggers status = Paid","","","High"),
    ("I","TC-133","Invoice","Employee cannot generate invoice","1. Login as Employee\n2. Open a Received PO","Generate Invoice button not visible to Employee role","","","High"),
    ("I","TC-134","Invoice","Invoice linked to correct PO","1. Generate invoice from PO #P-001\n2. Open invoice detail","Invoice shows reference to PO #P-001 with a working link","","","Medium"),
    ("I","TC-135","Invoice","Filter invoice list by payment status","1. Open Invoices list\n2. Filter by Paid","Only Paid invoices shown in list","","","Medium"),
    ("I","TC-136","Invoice","Invoice PDF/print generation","1. Open an invoice\n2. Click Download or Print","PDF generated with correct amounts, payment status, and customer details","","","Medium"),
    ("I","TC-137","Invoice","Paid invoice is read-only","1. Open invoice with status = Paid\n2. Try to modify any field or add a payment","All fields and actions locked. No modification possible","","","High"),
    ("I","TC-138","Invoice","Duplicate invoice prevention for same PO","1. Generate invoice from PO #P-001\n2. Try to generate another invoice from same PO","Blocked: An invoice already exists for this Purchase Order","","","High"),
    ("I","TC-139","Invoice","Generate invoice from PO not yet Received","1. Open a PO with status = Ordered\n2. Attempt to generate invoice","Invoice generation blocked: PO must be in Received status","","","High"),

    # ── SECTION J — Sales ─────────────────────────────────────────────────────
    ("J","TC-140","Sales","Create customer sales order with valid data","1. Navigate to Sales\n2. Create order for Customer A\n3. Add line items\n4. Save","Sales order saved with unique reference and correct status","","","High"),
    ("J","TC-141","Sales","Sales order linked to approved quotation","1. Approve a customer quotation\n2. Convert to Sales Order","Sales order references the originating quotation. Line items match","","","High"),
    ("J","TC-142","Sales","Sales order triggers job card creation","1. Create a Sales Order\n2. Click Create Job Card","Job card created with reference back to the sales order","","","High"),
    ("J","TC-143","Sales","Employee sees only assigned sales orders","1. Create 2 sales orders (1 to X, 1 to Y)\n2. Login as Employee X","Only Employee X sales order visible in their list","","","High"),
    ("J","TC-144","Sales","Manager sees all team sales orders","1. Sales orders assigned to Employee X and Y (both under Manager A)\n2. Login as Manager A","Both sales orders visible to Manager A","","","High"),
    ("J","TC-145","Sales","Update sales order status","1. Open a sales order\n2. Change status from In Progress to Dispatched","Status updates correctly. History log reflects the change with timestamp","","","Medium"),
    ("J","TC-146","Sales","Sales order total calculation","1. Add 2 line items with qty, price, and discount\n2. Check total","Total reflects correct arithmetic with all discounts applied","","","High"),
    ("J","TC-147","Sales","Cancel a sales order","1. Open an In Progress sales order\n2. Click Cancel","Status = Cancelled. Associated job card status also updated to Cancelled","","","Medium"),

    # ── SECTION K — Employee Management ──────────────────────────────────────
    ("K","TC-148","Employee Mgmt","Admin creates new employee user","1. Login as Admin\n2. Navigate to Users\n3. Create user role = Employee\n4. Assign to Manager A","User created. Appears in Manager A team immediately","","","High"),
    ("K","TC-149","Employee Mgmt","Admin creates new manager user","1. Login as Admin\n2. Create user with role = Manager","Manager account created. Can see team records once employees assigned","","","High"),
    ("K","TC-150","Employee Mgmt","Admin assigns employee to a manager","1. Login as Admin\n2. Edit Employee X\n3. Set reporting manager = Manager A","Employee X appears under Manager A team","","","High"),
    ("K","TC-151","Employee Mgmt","Admin reassigns employee to new manager","1. Employee X under Manager A\n2. Admin reassigns to Manager B","Manager B now sees Employee X records. Manager A no longer sees them","","","High"),
    ("K","TC-152","Employee Mgmt","Manager cannot create user accounts","1. Login as Manager\n2. Navigate to User Management","Create/Edit user actions not available to Manager role","","","High"),
    ("K","TC-153","Employee Mgmt","Admin upgrades employee role to Manager","1. Login as Admin\n2. Edit Employee X\n3. Change role to Manager","Employee X gets Manager-level access on next login","","","Medium"),
    ("K","TC-154","Employee Mgmt","Admin deactivates a user account","1. Login as Admin\n2. Deactivate Employee X","Employee X cannot login. Existing records remain visible to Manager/Admin","","","Medium"),
    ("K","TC-155","Employee Mgmt","Deactivated user cannot login","1. Admin deactivates Employee X\n2. Employee X attempts to login","Login rejected: Account deactivated or Invalid credentials","","","High"),
    ("K","TC-156","Employee Mgmt","Admin edits user profile details","1. Login as Admin\n2. Edit name/email of a user\n3. Save","Changes saved. Updated information reflects across the system","","","Medium"),
    ("K","TC-157","Employee Mgmt","Duplicate email on user creation","1. Login as Admin\n2. Create user with email already registered","Validation error: Email already registered","","","High"),

    # ── SECTION L — UI ────────────────────────────────────────────────────────
    ("L","TC-158","UI / Navigation","Sidebar navigation — all links work (all portals)","1. Login as each role type (Super Admin, Reseller, Admin, Manager, Employee)\n2. Click each sidebar menu item","Correct page loads for each role. No 404 errors encountered","","","High"),
    ("L","TC-159","UI / Navigation","Breadcrumb navigation accuracy","1. Navigate to a nested page\n2. Check breadcrumb","Breadcrumb shows correct path","","","Low"),
    ("L","TC-160","UI / Navigation","Back button preserves list filter state","1. Apply a filter on quotation list\n2. Open a record\n3. Press browser back","Returns to list with same filter still applied","","","Low"),
    ("L","TC-161","UI / Forms","Double-click submit prevention","1. Fill a quotation form completely\n2. Rapidly double-click Submit button","Only one record created. No duplicate quotation in the list","","","High"),
    ("L","TC-162","UI / Forms","Required field highlighting on submit","1. Open any create form\n2. Leave required fields blank\n3. Click Submit","All blank required fields highlighted in red with descriptive error text","","","High"),
    ("L","TC-163","UI / Forms","Form data persists on browser tab switch","1. Start filling a form\n2. Switch to another browser tab\n3. Return","Form data still intact. No data lost","","","Medium"),
    ("L","TC-164","UI / Forms","Long text input in description field","1. Enter 1000+ characters in any description field\n2. Save","Content accepted and saved. Displayed correctly on detail view","","","Low"),
    ("L","TC-165","UI / Forms","XSS injection in text fields","1. Enter script tag in a name or description field\n2. Save and reload","Input sanitized. Script does not execute. Stored as literal text","","","High"),
    ("L","TC-166","UI / Tables","Pagination on large dataset","1. Login as Admin\n2. Open list with 50+ records","Data loads in pages. Pagination controls work correctly","","","Medium"),
    ("L","TC-167","UI / Tables","Column sorting works correctly","1. Open any list table\n2. Click a column header","List re-orders ascending then descending on repeated clicks","","","Medium"),
    ("L","TC-168","UI / Tables","Export to Excel or CSV","1. Open any list\n2. Click Export button","File downloaded with correct column headers and all visible data","","","Medium"),
    ("L","TC-169","UI / Tables","Empty state shown for new user","1. Login as a brand new employee with no assigned records","List shows No records found message — not a blank white page","","","Medium"),

    # ── SECTION M — API ───────────────────────────────────────────────────────
    ("M","TC-170","API","GET /quotations with Admin token","1. Get Admin JWT via POST /auth/login\n2. Call GET /quotations","200 OK. Returns all quotations across all enterprise users","","","High"),
    ("M","TC-171","API","GET /quotations with Employee token","1. Get Employee X JWT\n2. Call GET /quotations","200 OK. Returns ONLY Employee X quotations","","","High"),
    ("M","TC-172","API","GET /quotations with Manager token","1. Get Manager A JWT\n2. Call GET /quotations","200 OK. Returns Manager A + all team members quotations","","","High"),
    ("M","TC-173","API","Reject quotation — version does not change","1. Get quotation at v1\n2. Call PATCH /quotations/:id/reject\n3. GET same quotation","Response shows version = 1 (unchanged after rejection)","","","High"),
    ("M","TC-174","API","Revise quotation — version increments","1. Get rejected quotation at v1\n2. Call PATCH /quotations/:id/revise\n3. GET same quotation","Response shows version = 2 after revision","","","High"),
    ("M","TC-175","API","POST /purchase-orders with Employee token","1. Get Employee JWT\n2. Call POST /purchase-orders","403 Forbidden returned","","","High"),
    ("M","TC-176","API","GET /purchase-orders scoped by employee","1. Get Employee X JWT\n2. Call GET /purchase-orders","Returns only POs where Employee X is the assigned owner","","","High"),
    ("M","TC-177","API","Super Admin endpoint with regular JWT","1. Login as Enterprise Admin\n2. Call GET /super-admin/enterprises with enterprise JWT","403 Forbidden — SuperAdminGuard blocks non-super-admin tokens","","","High"),
    ("M","TC-178","API","Reseller endpoint with Super Admin JWT","1. Login as Super Admin\n2. Call GET /resellers/me/tenants with super admin JWT","403 Forbidden — ResellerGuard blocks non-reseller tokens","","","High"),
    ("M","TC-179","API","All endpoints — missing Authorization header","1. Call any protected endpoint without Bearer token","401 Unauthorized response","","","High"),
    ("M","TC-180","API","All endpoints — invalid JWT token","1. Call endpoint with Authorization: Bearer invalidtokenxyz","401 Unauthorized response","","","High"),

    # ── SECTION N — Negative / Edge Cases ─────────────────────────────────────
    ("N","TC-181","Negative","Quotation with negative unit price","1. Create quotation\n2. Enter price = -500 on a line item\n3. Submit","Validation error: Price must be a positive value","","","Medium"),
    ("N","TC-182","Negative","PO created without selecting vendor","1. Start PO creation\n2. Leave vendor field blank\n3. Save","Validation error: Vendor is required","","","High"),
    ("N","TC-183","Negative","Invoice payment of 0","1. Open an invoice\n2. Record payment amount = 0","Validation error: Payment amount must be greater than 0","","","Medium"),
    ("N","TC-184","Negative","Edit approved quotation via direct API call","1. Get an Approved quotation ID\n2. Call PATCH /quotations/:id directly","400 or 403 error — must use /revise endpoint to modify","","","High"),
    ("N","TC-185","Negative","Concurrent approval of same quotation by two managers","1. Manager A and Manager B both open same Submitted quotation\n2. Both click Approve simultaneously","Only one approval registered. No duplicate status or conflicting state","","","Medium"),
]

current_section = None
current_row = 3
row_count = 0

for tc in test_cases:
    sec, tc_id, module, scenario, steps, expected, actual, status, priority = tc
    if sec != current_section:
        current_section = sec
        ws.merge_cells(f"A{current_row}:H{current_row}")
        sc = ws.cell(row=current_row, column=1, value=SECTION_LABELS[sec])
        sc.font = Font(name="Calibri", bold=True, color="F1F5F9", size=11, italic=True)
        sc.fill = fill("334155")
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sc.border = border()
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        row_count = 0

    row_bg = "FFFFFF" if row_count % 2 == 0 else "F8FAFC"
    row_count += 1
    values = [tc_id, module, scenario, steps, expected, actual, status, priority]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=current_row, column=col, value=val)
        c.font = Font(name="Calibri", bold=(col == 1), size=10)
        c.alignment = wrap_align("center" if col in (1, 7, 8) else "left")
        c.border = border()
        if col == 8 and priority in PRIORITY_FILL:
            c.fill = fill(PRIORITY_FILL[priority])
        elif col == 1:
            c.fill = fill("EFF6FF")
        else:
            c.fill = fill(row_bg)
    ws.row_dimensions[current_row].height = 70
    current_row += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{current_row - 1}"

# ── Summary Sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Coverage Summary")
for col, w in zip("ABCDE", [35, 14, 10, 10, 10]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells("A1:E1")
t = ws2["A1"]
t.value = "VABERP QA — Test Coverage Summary  |  v2.0"
t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t.fill = fill("1E293B")
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

for col, hdr in enumerate(["Module / Section", "Total TCs", "High", "Medium", "Low"], 1):
    c = ws2.cell(row=2, column=col, value=hdr)
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    c.fill = fill("334155")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
ws2.row_dimensions[2].height = 20

rows = [
    ("Authentication — All Roles", 18, 15, 3, 0),
    ("Super Admin Module", 22, 18, 4, 0),
    ("Reseller Admin Module", 18, 13, 4, 1),
    ("Role-Based Access Control", 9, 9, 0, 0),
    ("Manager Hierarchy", 6, 5, 1, 0),
    ("RFQ Module", 13, 7, 4, 2),
    ("Quotation Module", 24, 17, 6, 1),
    ("Purchase Order Module", 15, 11, 3, 1),
    ("Invoice Module", 14, 10, 4, 0),
    ("Sales Module", 8, 6, 2, 0),
    ("Employee Management", 10, 7, 3, 0),
    ("UI / Navigation / Forms", 12, 4, 5, 3),
    ("API Testing", 11, 11, 0, 0),
    ("Negative / Edge Cases", 5, 3, 2, 0),
]
for i, row in enumerate(rows):
    bg = "FFFFFF" if i % 2 == 0 else "F8FAFC"
    for col, val in enumerate(row, 1):
        c = ws2.cell(row=i+3, column=col, value=val)
        c.font = Font(name="Calibri", bold=(col==1), size=10)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center", indent=1 if col==1 else 0)
        c.fill = fill(bg)
        c.border = border()
    ws2.row_dimensions[i+3].height = 18

tr = len(rows) + 3
for col, val in enumerate(["TOTAL", 185, 136, 41, 8], 1):
    c = ws2.cell(row=tr, column=col, value=val)
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = fill("1E293B")
    c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center", indent=1 if col==1 else 0)
    c.border = border()
ws2.row_dimensions[tr].height = 20

# Role Hierarchy sheet
ws3 = wb.create_sheet("Role Hierarchy")
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 30
ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 20

ws3.merge_cells("A1:E1")
t3 = ws3["A1"]
t3.value = "VABERP — Complete Role Hierarchy & Permissions"
t3.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t3.fill = fill("1E293B")
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

for col, hdr in enumerate(["Role", "Portal / Login URL", "Scope", "Can Create", "Key Restrictions"], 1):
    c = ws3.cell(row=2, column=col, value=hdr)
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    c.fill = fill("334155")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border()
ws3.row_dimensions[2].height = 20

role_rows = [
    ("Super Admin", "/superadmin/login", "Full platform — all enterprises, all resellers, all data", "Enterprises, Resellers, Plans, Coupons, Services", "Cannot access ERP modules (Quotations, PO, Invoices)"),
    ("Reseller Admin", "/reseller/login", "Own tenant portfolio — only assigned enterprises", "Tenants, assign plans, set pricing", "Cannot access Super Admin portal or other resellers' tenants"),
    ("Enterprise Admin", "/login", "Own enterprise — all users, all modules, all data", "Users, Quotations, PO, Invoices, Sales", "Cannot access Super Admin or Reseller portals"),
    ("Manager", "/login", "Own records + direct report employees' records", "Quotations, RFQ, Sales Orders", "Cannot manage users. Cannot see other managers' team data"),
    ("Employee", "/login", "Only own assigned records", "RFQ, Quotations, Sales Orders (assigned only)", "Cannot approve, cannot create PO/Invoice, cannot manage users"),
]
role_colors = ["EDE9FE", "FEF3C7", "ECFDF5", "EFF6FF", "FFF1F2"]
for i, row in enumerate(role_rows):
    for col, val in enumerate(row, 1):
        c = ws3.cell(row=i+3, column=col, value=val)
        c.font = Font(name="Calibri", bold=(col==1), size=10)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        c.fill = fill(role_colors[i])
        c.border = border()
    ws3.row_dimensions[i+3].height = 45

# Credentials sheet
ws4 = wb.create_sheet("Test Credentials")
ws4.column_dimensions["A"].width = 25
ws4.column_dimensions["B"].width = 20
ws4.column_dimensions["C"].width = 32
ws4.column_dimensions["D"].width = 25
ws4.column_dimensions["E"].width = 18

ws4.merge_cells("A1:E1")
t4 = ws4["A1"]
t4.value = "VABERP — Test Environment Credentials"
t4.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t4.fill = fill("1E293B")
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 28

for col, hdr in enumerate(["Role", "Portal URL", "Email", "Password", "Notes"], 1):
    c = ws4.cell(row=2, column=col, value=hdr)
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    c.fill = fill("334155")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
ws4.row_dimensions[2].height = 20

cred_rows = [
    ("Super Admin", "http://64.235.43.187:2262/superadmin/login", "[SUPER_ADMIN_EMAIL]", "[SUPER_ADMIN_PASSWORD]", "Full platform access"),
    ("Reseller A", "http://64.235.43.187:2262/reseller/login", "[RESELLER_A_EMAIL]", "[RESELLER_A_PASSWORD]", "Has 5 test tenants"),
    ("Reseller B", "http://64.235.43.187:2262/reseller/login", "[RESELLER_B_EMAIL]", "[RESELLER_B_PASSWORD]", "For cross-reseller isolation tests"),
    ("Enterprise Admin", "http://64.235.43.187:2262/login", "[ADMIN_EMAIL]", "[ADMIN_PASSWORD]", "Full ERP access"),
    ("Manager A", "http://64.235.43.187:2262/login", "[MANAGER_A_EMAIL]", "[MANAGER_A_PASSWORD]", "Has Employee X and Y"),
    ("Manager B", "http://64.235.43.187:2262/login", "[MANAGER_B_EMAIL]", "[MANAGER_B_PASSWORD]", "Has Employee Z"),
    ("Employee X", "http://64.235.43.187:2262/login", "[EMP_X_EMAIL]", "[EMP_X_PASSWORD]", "Under Manager A"),
    ("Employee Y", "http://64.235.43.187:2262/login", "[EMP_Y_EMAIL]", "[EMP_Y_PASSWORD]", "Under Manager A"),
    ("Employee Z", "http://64.235.43.187:2262/login", "[EMP_Z_EMAIL]", "[EMP_Z_PASSWORD]", "Under Manager B"),
    ("API Base URL", "http://64.235.43.187:2261", "—", "—", "Use Bearer token from /auth/login, /superadmin/login, or /resellers/login"),
]
cred_colors = ["EDE9FE","FEF3C7","FEF3C7","ECFDF5","EFF6FF","EFF6FF","FFF1F2","FFF1F2","FFF1F2","F8FAFC"]
for i, row in enumerate(cred_rows):
    for col, val in enumerate(row, 1):
        c = ws4.cell(row=i+3, column=col, value=val)
        c.font = Font(name="Calibri", bold=(col==1), size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.fill = fill(cred_colors[i])
        c.border = border()
    ws4.row_dimensions[i+3].height = 20

ws4.cell(row=14, column=1, value="NOTE: Replace all [PLACEHOLDER] values with actual credentials before distributing to QA team.").font = Font(name="Calibri", italic=True, size=9, color="64748B")

out = r"C:\Users\UNITECH2\Desktop\VABERP_Test_Cases.xlsx"
wb.save(out)
print("Saved:", out)
