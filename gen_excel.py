from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Test Cases"

def side(): return Side(style="thin", color="CBD5E1")
def border(): return Border(left=side(), right=side(), top=side(), bottom=side())
def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def wrap_align(h="left"): return Alignment(wrap_text=True, vertical="top", horizontal=h)

COLUMNS = [
    ("TC ID", 12), ("Module", 18), ("Test Scenario", 38),
    ("Test Steps", 55), ("Expected Result", 50),
    ("Actual Result", 28), ("Status", 12), ("Priority", 12),
]
for i, (_, w) in enumerate(COLUMNS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title
ws.merge_cells("A1:H1")
t = ws["A1"]
t.value = "VABERP — Complete QA Test Cases  |  145 Test Cases  |  v1.0  |  2026-04-07"
t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t.fill = fill("1E293B")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Header row
for col, (name, _) in enumerate(COLUMNS, 1):
    c = ws.cell(row=2, column=col, value=name)
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    c.fill = fill("334155")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border()
ws.row_dimensions[2].height = 22

SECTION_LABELS = {
    "A": "SECTION A — Authentication (TC-001 to TC-013)",
    "B": "SECTION B — Role-Based Access Control (TC-014 to TC-022)",
    "C": "SECTION C — Manager Hierarchy & Visibility (TC-023 to TC-028)",
    "D": "SECTION D — RFQ Module (TC-029 to TC-041)",
    "E": "SECTION E — Quotation Module (TC-042 to TC-065)",
    "F": "SECTION F — Purchase Order Module (TC-066 to TC-080)",
    "G": "SECTION G — Invoice Module (TC-081 to TC-094)",
    "H": "SECTION H — Sales Module (TC-095 to TC-102)",
    "I": "SECTION I — Employee & Role Management (TC-103 to TC-112)",
    "J": "SECTION J — UI, Navigation & Forms (TC-113 to TC-124)",
    "K": "SECTION K — API Testing (TC-125 to TC-135)",
    "L": "SECTION L — Negative & Edge Cases (TC-136 to TC-145)",
}

PRIORITY_FILL = {"High": "FEE2E2", "Medium": "FEF9C3", "Low": "DCFCE7"}

test_cases = [
    ("A","TC-001","Auth","Admin login with valid credentials","1. Go to /login\n2. Enter admin email & password\n3. Click Login","Admin dashboard loads with full navigation menu visible","","","High"),
    ("A","TC-002","Auth","Manager login with valid credentials","1. Go to /login\n2. Enter manager credentials\n3. Click Login","Manager dashboard loads; team-scoped menu items visible","","","High"),
    ("A","TC-003","Auth","Employee login with valid credentials","1. Go to /login\n2. Enter employee credentials\n3. Click Login","Employee dashboard loads; restricted menu (no PO create, no approvals)","","","High"),
    ("A","TC-004","Auth","Login with invalid password","1. Enter valid email + wrong password\n2. Click Login","Error: Invalid credentials. No redirect. Token not issued","","","High"),
    ("A","TC-005","Auth","Login with unregistered email","1. Enter email not in system\n2. Enter any password\n3. Click Login","Error: User not found or Invalid credentials","","","High"),
    ("A","TC-006","Auth","Login with blank email field","1. Leave email blank\n2. Fill password\n3. Click Login","Frontend validation: Email is required","","","Medium"),
    ("A","TC-007","Auth","Login with blank password field","1. Fill email\n2. Leave password blank\n3. Click Login","Frontend validation: Password is required","","","Medium"),
    ("A","TC-008","Auth","Login with both fields blank","1. Leave both fields blank\n2. Click Login","Both field validation errors shown. No API call made","","","Medium"),
    ("A","TC-009","Auth","Session persistence after page refresh","1. Login as any user\n2. Refresh browser","User remains logged in. Dashboard reloads without re-authentication","","","High"),
    ("A","TC-010","Auth","Logout clears session","1. Login as any user\n2. Click Logout","Session cleared. Redirected to /login. Back button does not return to dashboard","","","High"),
    ("A","TC-011","Auth","Expired session redirects to login","1. Login\n2. Manually clear auth token from browser storage\n3. Try to navigate","Redirected to /login. Unauthorized message shown","","","High"),
    ("A","TC-012","Auth","Brute force — repeated wrong password","1. Enter wrong password 5 times consecutively","Account locked or rate-limit error message shown","","","Medium"),
    ("A","TC-013","Auth","SQL injection in login email field","1. Enter ' OR 1=1 -- in email field\n2. Click Login","Login rejected. No data exposed. Standard error message shown","","","High"),
    ("B","TC-014","RBAC","Employee cannot access PO creation","1. Login as Employee\n2. Navigate to /purchase-orders/new","Access denied or button not visible. No PO form rendered","","","High"),
    ("B","TC-015","RBAC","Employee cannot generate invoices","1. Login as Employee\n2. Open a Received PO\n3. Check for Generate Invoice action","Generate Invoice button not visible to Employee","","","High"),
    ("B","TC-016","RBAC","Employee cannot access User Management","1. Login as Employee\n2. Navigate to /settings/users via URL","Redirect to dashboard or 403 error. No user list shown","","","High"),
    ("B","TC-017","RBAC","Manager cannot access User Management","1. Login as Manager\n2. Navigate to /settings/users","Access denied. Create/Edit user actions unavailable","","","High"),
    ("B","TC-018","RBAC","Manager can approve quotations","1. Login as Manager\n2. Open Submitted quotation\n3. Click Approve","Approval action succeeds. Status changes to Approved","","","High"),
    ("B","TC-019","RBAC","Employee cannot approve quotations","1. Login as Employee\n2. Open a Submitted quotation","Approve button not visible. Cannot trigger approval action","","","High"),
    ("B","TC-020","RBAC","Admin can access all modules","1. Login as Admin\n2. Navigate through every menu item","All modules accessible without restriction","","","High"),
    ("B","TC-021","RBAC","Direct URL access by unauthorized role","1. Login as Employee\n2. Manually type /purchase-orders/new in URL bar","Redirected away or 403 error shown. No form rendered","","","High"),
    ("B","TC-022","RBAC","API call with Employee token to Manager endpoint","1. Login as Employee via Postman\n2. Call POST /purchase-orders with employee JWT","403 Forbidden response returned","","","High"),
    ("C","TC-023","Hierarchy","Manager A sees Employee X and Y records","1. Employee X and Y (both under Manager A) each create a quotation\n2. Login as Manager A\n3. Open Quotations list","Records of both Employee X and Employee Y visible to Manager A","","","High"),
    ("C","TC-024","Hierarchy","Manager A cannot see Employee Z records","1. Employee Z (under Manager B) creates a quotation\n2. Login as Manager A\n3. Open Quotations list","Employee Z quotation NOT in Manager A list","","","High"),
    ("C","TC-025","Hierarchy","Manager B cannot see Manager A team records","1. Employee X (under Manager A) creates a record\n2. Login as Manager B","Employee X record not visible to Manager B","","","High"),
    ("C","TC-026","Hierarchy","Admin sees all records from all users","1. Employees under Manager A and B each create records\n2. Login as Admin","All records from all users visible","","","High"),
    ("C","TC-027","Hierarchy","Employee reassigned — visibility updates immediately","1. Employee X is under Manager A\n2. Admin reassigns Employee X to Manager B\n3. Login as Manager B","Employee X records now visible to Manager B. No longer visible to Manager A","","","High"),
    ("C","TC-028","Hierarchy","Manager own records visible to themselves","1. Manager A creates a quotation directly\n2. Login as Manager A","Manager A own quotation visible in their list","","","Medium"),
    ("D","TC-029","RFQ","Create RFQ with valid data","1. Login\n2. Navigate to RFQ\n3. Fill vendor, items, quantities\n4. Submit","RFQ created with status Draft/Sent and unique reference number","","","High"),
    ("D","TC-030","RFQ","Create RFQ with missing vendor","1. Fill RFQ items\n2. Leave vendor blank\n3. Submit","Validation error: Vendor is required","","","High"),
    ("D","TC-031","RFQ","Create RFQ with zero quantity","1. Create RFQ\n2. Set item qty = 0\n3. Submit","Validation error: Quantity must be greater than 0","","","Medium"),
    ("D","TC-032","RFQ","Create RFQ with negative quantity","1. Create RFQ\n2. Enter qty = -5\n3. Submit","Validation error: negative values not allowed","","","Medium"),
    ("D","TC-033","RFQ","Create RFQ with no line items","1. Fill vendor\n2. Add no items\n3. Submit","Validation error: At least one line item is required","","","High"),
    ("D","TC-034","RFQ","Convert RFQ to Quotation","1. Open a submitted/received RFQ\n2. Click Convert to Quotation","Quotation created at v1 with all items pre-filled from RFQ","","","High"),
    ("D","TC-035","RFQ","Quotation from RFQ pre-fills all line items","1. Create RFQ with 3 items\n2. Convert to Quotation\n3. Open the quotation","All 3 items present in quotation with matching quantity and description","","","High"),
    ("D","TC-036","RFQ","Employee sees only their own RFQs","1. Employee X creates an RFQ\n2. Login as Employee Y","Employee Y cannot see Employee X RFQ","","","High"),
    ("D","TC-037","RFQ","Manager sees all team RFQs","1. Employee X (under Manager A) creates an RFQ\n2. Login as Manager A","RFQ appears in Manager A list","","","High"),
    ("D","TC-038","RFQ","Duplicate RFQ same vendor same reference","1. Create RFQ for Vendor A ref #001\n2. Create another with same vendor and reference","System warns or blocks exact duplicate","","","Medium"),
    ("D","TC-039","RFQ","RFQ list displays correct status badges","1. Create RFQs in different states\n2. Open RFQ list","Status badges correctly show Draft, Sent, Received per record","","","Low"),
    ("D","TC-040","RFQ","Delete a Draft RFQ","1. Create an RFQ\n2. Keep it in Draft\n3. Delete it","RFQ removed from list. No linked records affected","","","Medium"),
    ("D","TC-041","RFQ","Cannot delete a converted RFQ","1. Convert RFQ to Quotation\n2. Try to delete the original RFQ","Deletion blocked: RFQ has a linked Quotation","","","Medium"),
    ("E","TC-042","Quotation","Create quotation with valid line items","1. Navigate to Create Quotation\n2. Add items with qty, unit price, discount\n3. Save","Quotation saved at v1, status = Draft, total calculated correctly","","","High"),
    ("E","TC-043","Quotation","Total calculation — single item with discount","1. Add Item: qty=2, price=1000, discount=10%","Total = (2x1000) - 10% = 1800","","","High"),
    ("E","TC-044","Quotation","Total calculation — multiple items with discounts","1. Item A: qty=1, price=500, 0% discount\n2. Item B: qty=2, price=300, 5% discount","Total = 500 + 570 = 1070","","","High"),
    ("E","TC-045","Quotation","Discount exactly 100%","1. Create quotation\n2. Enter discount = 100% on a line item","Accepted — line item cost = 0. Total reflects correctly","","","Medium"),
    ("E","TC-046","Quotation","Discount over 100%","1. Create quotation\n2. Enter discount = 101%\n3. Save","Validation error: Discount cannot exceed 100%","","","High"),
    ("E","TC-047","Quotation","Discount = 0% no discount","1. Leave discount blank or enter 0 on a line item","No discount applied. Full price used in total","","","Low"),
    ("E","TC-048","Quotation","Create quotation with no line items","1. Open Create Quotation\n2. Add no line items\n3. Submit","Validation error: At least one line item is required","","","High"),
    ("E","TC-049","Quotation","Submit quotation for approval","1. Open a Draft quotation\n2. Click Submit","Status = Submitted. Approve/Reject options now available to approvers","","","High"),
    ("E","TC-050","Quotation","Approve a submitted quotation","1. Login as Manager or Admin\n2. Open Submitted quotation\n3. Click Approve","Status = Approved. Version number unchanged. Edit fields locked","","","High"),
    ("E","TC-051","Quotation","Reject a submitted quotation","1. Login as Manager or Admin\n2. Open Submitted quotation\n3. Click Reject\n4. Enter rejection reason","Status = Rejected. Version number stays the same (e.g. v1 stays v1)","","","High"),
    ("E","TC-052","Quotation","CRITICAL: Version does NOT increment on rejection","1. Create Quotation (version = 1)\n2. Submit\n3. Manager rejects\n4. Check version field","Version field = 1 unchanged. Rejection must NEVER auto-increment version","","","High"),
    ("E","TC-053","Quotation","Revise a rejected quotation","1. Open Rejected quotation\n2. Click Revise\n3. Modify at least one line item\n4. Resubmit","New version created: v1 to v2. Original v1 preserved in version history","","","High"),
    ("E","TC-054","Quotation","CRITICAL: Version increments on user revision","1. Reject a v1 quotation\n2. User clicks Revise\n3. Check version after save","Version = 2. Re-submission shows v2 to approver","","","High"),
    ("E","TC-055","Quotation","Multiple revision cycles — correct versioning","1. v1 Reject Revise v2 Reject Revise v3","Versions: v1, v2, v3 in correct sequence. All listed in history","","","High"),
    ("E","TC-056","Quotation","Version history shows all revisions with metadata","1. Create quotation with 3 revision cycles\n2. Open Version History tab","v1, v2, v3 listed with timestamp, status, and who approved or rejected","","","Medium"),
    ("E","TC-057","Quotation","Approved quotation fields are read-only","1. Open an Approved quotation\n2. Try to click and edit any field","All fields read-only. Only Revise action button available","","","High"),
    ("E","TC-058","Quotation","Revise approved quotation creates new version","1. Open Approved quotation\n2. Click Revise\n3. Edit and resubmit","New version (e.g. v2) created in Draft. Previous approved version preserved","","","High"),
    ("E","TC-059","Quotation","Cannot re-submit already Submitted quotation","1. Submit a quotation (status = Submitted)\n2. Click Submit again","Submit button disabled or already-submitted message shown. No duplicate submission","","","Medium"),
    ("E","TC-060","Quotation","Employee cannot approve own submitted quotation","1. Login as Employee\n2. Create and submit a quotation\n3. Check for Approve button","Approve button NOT visible to the Employee who submitted it","","","High"),
    ("E","TC-061","Quotation","Quotation linked to correct customer","1. Create quotation for Customer A\n2. Open quotation detail","Customer A shown in customer field. No data leakage from other customers","","","High"),
    ("E","TC-062","Quotation","Search quotation by customer name","1. Open Quotation list\n2. Search by customer name","Only quotations linked to that customer are shown","","","Medium"),
    ("E","TC-063","Quotation","Filter quotations by status","1. Open Quotation list\n2. Apply filter Approved","Only Approved quotations displayed in list","","","Medium"),
    ("E","TC-064","Quotation","Quotation PDF/print generation","1. Open an Approved quotation\n2. Click Download or Print","PDF generated with correct line items, totals, customer details, and version","","","Medium"),
    ("E","TC-065","Quotation","Duplicate quotation reference prevention","1. Create Quotation ref #Q-001 for Customer A\n2. Attempt to create another with same ref/customer","System warns or prevents exact duplicate quotation creation","","","Medium"),
    ("F","TC-066","Purchase Order","Create PO from approved quotation","1. Open Approved quotation\n2. Click Create PO\n3. Assign to Employee\n4. Save","PO created with correct amounts and assigned to selected employee","","","High"),
    ("F","TC-067","Purchase Order","PO amount matches approved quotation total","1. Approved quotation total = 25000\n2. Create PO from it","PO total = 25000. No discrepancy between quotation and PO","","","High"),
    ("F","TC-068","Purchase Order","Cannot create PO from Draft quotation","1. Open a Draft quotation\n2. Check for Create PO option","Create PO button not visible or disabled","","","High"),
    ("F","TC-069","Purchase Order","Cannot create PO from Rejected quotation","1. Open a Rejected quotation\n2. Check for Create PO option","Create PO button not available for rejected quotations","","","High"),
    ("F","TC-070","Purchase Order","Cannot create PO from Submitted quotation","1. Open a Submitted (pending approval) quotation\n2. Check Create PO","Create PO disabled — must wait for approval","","","High"),
    ("F","TC-071","Purchase Order","Duplicate PO prevention for same quotation","1. Create PO from Approved quotation\n2. Try to create another PO from same quotation","System blocks or warns: A PO already exists for this quotation","","","High"),
    ("F","TC-072","Purchase Order","Employee sees only their assigned POs","1. Create PO assigned to Employee X\n2. Login as Employee X\n3. Open PO list","PO assigned to Employee X is visible","","","High"),
    ("F","TC-073","Purchase Order","Employee cannot see PO assigned to another employee","1. Create PO assigned to Employee Y\n2. Login as Employee X\n3. Open PO list","Employee Y PO is NOT in Employee X list","","","High"),
    ("F","TC-074","Purchase Order","Manager sees all team POs","1. Create POs for Employee X and Y (both under Manager A)\n2. Login as Manager A","Both POs from Employee X and Y visible to Manager A","","","High"),
    ("F","TC-075","Purchase Order","Update PO status to Received","1. Open an active PO\n2. Update status to Received","Status updates to Received. Invoice generation now enabled","","","High"),
    ("F","TC-076","Purchase Order","PO cannot be deleted if linked invoice exists","1. Generate an invoice from PO\n2. Try to delete the PO","Deletion blocked: Cannot delete a PO with an associated invoice","","","High"),
    ("F","TC-077","Purchase Order","PO list sorted by date newest first","1. Create 3 POs on different dates\n2. Open PO list","Most recently created PO appears first","","","Low"),
    ("F","TC-078","Purchase Order","Search PO by vendor name","1. Open PO list\n2. Search Vendor A","Only POs linked to Vendor A shown","","","Medium"),
    ("F","TC-079","Purchase Order","PO reference number is unique","1. Create multiple POs","Each PO has a unique system-generated reference number","","","High"),
    ("F","TC-080","Purchase Order","Reassign PO updates visibility","1. PO assigned to Employee X\n2. Edit PO and reassign to Employee Y","PO now visible to Employee Y. No longer visible to Employee X","","","Medium"),
    ("G","TC-081","Invoice","Generate invoice from Received PO","1. Open PO with status = Received\n2. Click Generate Invoice","Invoice created with amount pre-filled from PO. Status = Unpaid","","","High"),
    ("G","TC-082","Invoice","Invoice amount auto-populated from PO","1. PO total = 40000\n2. Generate invoice","Invoice amount = 40000. No manual entry required","","","High"),
    ("G","TC-083","Invoice","Invoice default status is Unpaid","1. Generate any new invoice","Invoice status = Unpaid on creation","","","High"),
    ("G","TC-084","Invoice","Record partial payment","1. Invoice total = 50000\n2. Record payment of 20000","Status = Partially Paid. Remaining balance = 30000 shown","","","High"),
    ("G","TC-085","Invoice","Record full payment","1. Invoice total = 50000\n2. Record payment of 50000","Status = Paid. Balance = 0. Invoice marked complete","","","High"),
    ("G","TC-086","Invoice","Payment amount exceeds invoice total","1. Invoice = 50000\n2. Attempt to record 60000 payment","Validation error: Payment cannot exceed invoice total","","","High"),
    ("G","TC-087","Invoice","Multiple partial payments until fully paid","1. Invoice = 50000\n2. Record 20000 then 20000 then 10000","Each payment recorded. Final payment triggers status = Paid. All records listed","","","High"),
    ("G","TC-088","Invoice","Employee cannot generate invoice","1. Login as Employee\n2. Open a Received PO","Generate Invoice button not visible to Employee role","","","High"),
    ("G","TC-089","Invoice","Invoice linked to correct PO with clickable reference","1. Generate invoice from PO #P-001\n2. Open invoice detail","Invoice shows reference to PO #P-001 with a working link","","","Medium"),
    ("G","TC-090","Invoice","Navigate from invoice back to source PO","1. Open an invoice\n2. Click PO reference link","Correctly navigates to the originating PO record","","","Medium"),
    ("G","TC-091","Invoice","Filter invoice list by payment status","1. Open Invoices list\n2. Filter by Paid","Only Paid invoices shown in list","","","Medium"),
    ("G","TC-092","Invoice","Invoice PDF/print generation","1. Open an invoice\n2. Click Download or Print","PDF generated with correct amounts, payment status, and customer details","","","Medium"),
    ("G","TC-093","Invoice","Paid invoice is read-only","1. Open invoice with status = Paid\n2. Try to modify any field or add a payment","All fields and actions locked. No modification possible","","","High"),
    ("G","TC-094","Invoice","Duplicate invoice prevention for same PO","1. Generate invoice from PO #P-001\n2. Try to generate another invoice from same PO","Blocked: An invoice already exists for this Purchase Order","","","High"),
    ("H","TC-095","Sales","Create customer sales order with valid data","1. Navigate to Sales\n2. Create order for Customer A\n3. Add line items\n4. Save","Sales order saved with unique reference and correct status","","","High"),
    ("H","TC-096","Sales","Sales order linked to approved quotation","1. Approve a customer quotation\n2. Convert to Sales Order","Sales order references the originating quotation. Line items match","","","High"),
    ("H","TC-097","Sales","Sales order triggers job card creation","1. Create a Sales Order\n2. Click Create Job Card","Job card created with reference back to the sales order","","","High"),
    ("H","TC-098","Sales","Employee sees only assigned sales orders","1. Create 2 sales orders (1 assigned to X, 1 to Y)\n2. Login as Employee X","Only Employee X sales order visible in their list","","","High"),
    ("H","TC-099","Sales","Manager sees all team sales orders","1. Sales orders assigned to Employee X and Y (both under Manager A)\n2. Login as Manager A","Both sales orders visible to Manager A","","","High"),
    ("H","TC-100","Sales","Update sales order status","1. Open a sales order\n2. Change status from In Progress to Dispatched","Status updates correctly. History log reflects the change with timestamp","","","Medium"),
    ("H","TC-101","Sales","Sales order total calculation","1. Add 2 line items with qty, price, and discount\n2. Check total","Total reflects correct arithmetic with all discounts applied","","","High"),
    ("H","TC-102","Sales","Cancel a sales order","1. Open an In Progress sales order\n2. Click Cancel","Status = Cancelled. Associated job card status also updated to Cancelled","","","Medium"),
    ("I","TC-103","Employee Mgmt","Admin creates new employee user","1. Login as Admin\n2. Navigate to Users\n3. Create user role = Employee\n4. Assign to Manager A","User created. Appears in Manager A team immediately","","","High"),
    ("I","TC-104","Employee Mgmt","Admin creates new manager user","1. Login as Admin\n2. Create user with role = Manager","Manager account created. Can see team records once employees assigned","","","High"),
    ("I","TC-105","Employee Mgmt","Admin assigns employee to a manager","1. Login as Admin\n2. Edit Employee X\n3. Set reporting manager = Manager A","Employee X appears under Manager A team","","","High"),
    ("I","TC-106","Employee Mgmt","Admin reassigns employee to new manager","1. Employee X under Manager A\n2. Admin reassigns to Manager B","Manager B now sees Employee X records. Manager A no longer sees them","","","High"),
    ("I","TC-107","Employee Mgmt","Manager cannot create user accounts","1. Login as Manager\n2. Navigate to User Management","Create/Edit user actions not available to Manager role","","","High"),
    ("I","TC-108","Employee Mgmt","Admin upgrades employee role to Manager","1. Login as Admin\n2. Edit Employee X\n3. Change role to Manager","Employee X gets Manager-level access on next login","","","Medium"),
    ("I","TC-109","Employee Mgmt","Admin deactivates a user account","1. Login as Admin\n2. Deactivate Employee X","Employee X cannot login. Their existing records remain visible to Manager/Admin","","","Medium"),
    ("I","TC-110","Employee Mgmt","Deactivated user cannot login","1. Admin deactivates Employee X\n2. Employee X attempts to login with correct credentials","Login rejected: Account deactivated or Invalid credentials","","","High"),
    ("I","TC-111","Employee Mgmt","Admin edits user profile details","1. Login as Admin\n2. Edit name/email of a user\n3. Save","Changes saved. Updated information reflects across the system","","","Medium"),
    ("I","TC-112","Employee Mgmt","Duplicate email on user creation","1. Login as Admin\n2. Create user with email already registered in system","Validation error: Email already registered","","","High"),
    ("J","TC-113","UI / Navigation","Sidebar navigation — all links work","1. Login as Admin\n2. Click each sidebar menu item","Correct page loads for each. No 404 errors encountered","","","High"),
    ("J","TC-114","UI / Navigation","Breadcrumb navigation accuracy","1. Navigate to a nested page (e.g. Quotation detail)\n2. Check breadcrumb","Breadcrumb shows correct path: Home > Quotations > #Q-001","","","Low"),
    ("J","TC-115","UI / Navigation","Back button preserves list filter state","1. Apply a filter on quotation list\n2. Open a record\n3. Press browser back","Returns to list with same filter still applied","","","Low"),
    ("J","TC-116","UI / Forms","Double-click submit prevention","1. Fill a quotation form completely\n2. Rapidly double-click Submit button","Only one record created. No duplicate quotation in the list","","","High"),
    ("J","TC-117","UI / Forms","Required field highlighting on submit","1. Open any create form\n2. Leave required fields blank\n3. Click Submit","All blank required fields highlighted in red with descriptive error text","","","High"),
    ("J","TC-118","UI / Forms","Form data persists on browser tab switch","1. Start filling a form\n2. Switch to another browser tab\n3. Return","Form data still intact. No data lost","","","Medium"),
    ("J","TC-119","UI / Forms","Long text input in description field","1. Enter 1000+ characters in any description field\n2. Save","Content accepted and saved. Displayed correctly on detail view","","","Low"),
    ("J","TC-120","UI / Forms","XSS injection in text fields","1. Enter script tag in a name or description field\n2. Save and reload","Input sanitized. Script does not execute. Stored and displayed as literal text","","","High"),
    ("J","TC-121","UI / Tables","Pagination on large dataset","1. Login as Admin\n2. Open list with 50+ records","Data loads in pages. Pagination controls work correctly","","","Medium"),
    ("J","TC-122","UI / Tables","Column sorting works correctly","1. Open any list table\n2. Click a column header","List re-orders ascending then descending on repeated clicks","","","Medium"),
    ("J","TC-123","UI / Tables","Export to Excel or CSV","1. Open any list\n2. Click Export button","File downloaded with correct column headers and all visible data","","","Medium"),
    ("J","TC-124","UI / Tables","Empty state shown for new user","1. Login as a brand new employee with no assigned records","List shows No records found message — not a blank white page","","","Medium"),
    ("K","TC-125","API","GET /quotations with Admin token","1. Get Admin JWT via POST /auth/login\n2. Call GET /quotations","200 OK. Returns all quotations across all users","","","High"),
    ("K","TC-126","API","GET /quotations with Employee token","1. Get Employee X JWT\n2. Call GET /quotations","200 OK. Returns ONLY Employee X quotations. No other user data","","","High"),
    ("K","TC-127","API","GET /quotations with Manager token","1. Get Manager A JWT\n2. Call GET /quotations","200 OK. Returns Manager A + all team members quotations","","","High"),
    ("K","TC-128","API","POST /quotations with missing required field","1. Call POST /quotations without customer_id in body","400 Bad Request with descriptive validation message","","","High"),
    ("K","TC-129","API","Reject quotation — version does not change","1. Get quotation at v1\n2. Call PATCH /quotations/:id/reject\n3. GET same quotation","Response shows version = 1 (unchanged after rejection)","","","High"),
    ("K","TC-130","API","Revise quotation — version increments","1. Get rejected quotation at v1\n2. Call PATCH /quotations/:id/revise\n3. GET same quotation","Response shows version = 2 after revision","","","High"),
    ("K","TC-131","API","POST /purchase-orders with Employee token","1. Get Employee JWT\n2. Call POST /purchase-orders","403 Forbidden returned","","","High"),
    ("K","TC-132","API","GET /purchase-orders scoped by employee","1. Get Employee X JWT\n2. Call GET /purchase-orders","Returns only POs where Employee X is the assigned owner","","","High"),
    ("K","TC-133","API","POST /invoices — duplicate for same PO","1. Generate invoice for PO #P-001\n2. Call POST /invoices again for same PO ID","409 Conflict or 400 validation error returned","","","High"),
    ("K","TC-134","API","All endpoints — missing Authorization header","1. Call any protected endpoint without Bearer token","401 Unauthorized response","","","High"),
    ("K","TC-135","API","All endpoints — invalid JWT token","1. Call endpoint with Authorization: Bearer invalidtokenxyz","401 Unauthorized response","","","High"),
    ("L","TC-136","Negative","Quotation with negative unit price","1. Create quotation\n2. Enter price = -500 on a line item\n3. Submit","Validation error: Price must be a positive value","","","Medium"),
    ("L","TC-137","Negative","PO created without selecting vendor","1. Start PO creation\n2. Leave vendor field blank\n3. Save","Validation error: Vendor is required","","","High"),
    ("L","TC-138","Negative","Invoice payment of 0","1. Open an invoice\n2. Record payment amount = 0","Validation error: Payment amount must be greater than 0","","","Medium"),
    ("L","TC-139","Negative","Navigate to URL of deleted record","1. Delete a quotation\n2. Navigate to its direct URL","404 Not Found or Record not found message shown","","","Medium"),
    ("L","TC-140","Negative","Edit approved quotation via direct API call","1. Get an Approved quotation ID\n2. Call PATCH /quotations/:id directly with field modifications","400 or 403 error — must use /revise endpoint to modify","","","High"),
    ("L","TC-141","Negative","Generate invoice from PO not yet Received","1. Open a PO with status = Ordered (not Received)\n2. Attempt to generate invoice","Invoice generation blocked: PO must be in Received status","","","High"),
    ("L","TC-142","Negative","Delete employee who has open active records","1. Employee X has open quotations and POs\n2. Admin tries to delete Employee X","Deletion blocked or prompt to reassign records shown","","","High"),
    ("L","TC-143","Negative","Assign PO to non-existent employee via API","1. Call POST /purchase-orders with employee_id = 99999","400 or 404 — employee not found error","","","Medium"),
    ("L","TC-144","Negative","XSS in quotation description via API","1. Call POST /quotations with XSS payload in description\n2. GET the record","Script not executed on retrieval. Stored and returned as sanitized plain string","","","High"),
    ("L","TC-145","Negative","Concurrent approval of same quotation by two managers","1. Manager A and Manager B both open same Submitted quotation\n2. Both click Approve simultaneously","Only one approval registered. No duplicate status or conflicting state","","","Medium"),
]

current_row = 3
current_section = None
row_count = 0

for tc in test_cases:
    sec, tc_id, module, scenario, steps, expected, actual, status, priority = tc
    if sec != current_section:
        current_section = sec
        ws.merge_cells(f"A{current_row}:H{current_row}")
        sc = ws.cell(row=current_row, column=1, value=SECTION_LABELS[sec])
        sc.font = Font(name="Calibri", bold=True, color="F1F5F9", size=11, italic=True)
        sc.fill = fill("334155")
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sc.border = border()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    row_bg = "FFFFFF" if row_count % 2 == 0 else "F8FAFC"
    row_count += 1
    values = [tc_id, module, scenario, steps, expected, actual, status, priority]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=current_row, column=col, value=val)
        c.font = Font(name="Calibri", bold=(col == 1), size=10)
        c.alignment = wrap_align("center" if col in (1, 7, 8) else "left")
        c.border = border()
        if col == 8 and priority in PRIORITY_FILL:
            c.fill = fill(PRIORITY_FILL[priority])
        elif col == 1:
            c.fill = fill("EFF6FF")
        else:
            c.fill = fill(row_bg)
    ws.row_dimensions[current_row].height = 70
    current_row += 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{current_row - 1}"

# ── Summary Sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Coverage Summary")
for col, w in zip("ABCDE", [32, 14, 10, 10, 10]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells("A1:E1")
t = ws2["A1"]
t.value = "VABERP QA — Test Coverage Summary"
t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
t.fill = fill("1E293B")
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

for col, hdr in enumerate(["Module / Section", "Total TCs", "High", "Medium", "Low"], 1):
    c = ws2.cell(row=2, column=col, value=hdr)
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    c.fill = fill("334155")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
ws2.row_dimensions[2].height = 20

rows = [
    ("Authentication", 13, 9, 3, 1),
    ("Role-Based Access Control", 9, 9, 0, 0),
    ("Manager Hierarchy", 6, 5, 1, 0),
    ("RFQ Module", 13, 7, 4, 2),
    ("Quotation Module", 24, 17, 6, 1),
    ("Purchase Order Module", 15, 11, 3, 1),
    ("Invoice Module", 14, 10, 4, 0),
    ("Sales Module", 8, 6, 2, 0),
    ("Employee Management", 10, 7, 3, 0),
    ("UI / Navigation / Forms", 12, 4, 5, 3),
    ("API Testing", 11, 11, 0, 0),
    ("Negative / Edge Cases", 10, 6, 4, 0),
]
for i, row in enumerate(rows):
    bg = "FFFFFF" if i % 2 == 0 else "F8FAFC"
    for col, val in enumerate(row, 1):
        c = ws2.cell(row=i+3, column=col, value=val)
        c.font = Font(name="Calibri", bold=(col==1), size=10)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center", indent=1 if col==1 else 0)
        c.fill = fill(bg)
        c.border = border()
    ws2.row_dimensions[i+3].height = 18

tr = len(rows) + 3
for col, val in enumerate(["TOTAL", 145, 102, 35, 8], 1):
    c = ws2.cell(row=tr, column=col, value=val)
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = fill("1E293B")
    c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center", indent=1 if col==1 else 0)
    c.border = border()
ws2.row_dimensions[tr].height = 20

ws2.cell(row=tr+2, column=1, value="Priority Legend").font = Font(name="Calibri", bold=True, size=11)
for j, (p, desc, col) in enumerate([
    ("High", "Business-critical. Blocks release if failed.", "FEE2E2"),
    ("Medium", "Important, should pass before release.", "FEF9C3"),
    ("Low", "Minor / cosmetic. Can be deferred.", "DCFCE7"),
]):
    r = tr + 3 + j
    pc = ws2.cell(row=r, column=1, value=p)
    pc.fill = fill(col); pc.font = Font(name="Calibri", bold=True, size=10); pc.border = border()
    ws2.merge_cells(f"B{r}:E{r}")
    dc = ws2.cell(row=r, column=2, value=desc)
    dc.font = Font(name="Calibri", size=10); dc.border = border()
    ws2.row_dimensions[r].height = 16

out = r"C:\Users\UNITECH2\Desktop\VABERP_Test_Cases.xlsx"
wb.save(out)
print("Saved:", out)
