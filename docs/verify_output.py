"""Quick verification of the locked-fallback xlsx contents."""
from openpyxl import load_workbook
import os

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "ERP_Enterprise.locked.xlsx")

wb = load_workbook(XLSX)
print("Sheets:", wb.sheetnames)

# Stats sheet
print("\n=== Test Run Stats sheet ===")
ws = wb["Test Run Stats"]
for r in range(1, 20):
    row = [ws.cell(row=r, column=c).value for c in range(1, 5)]
    if any(v is not None for v in row):
        print(f"  row {r}:", row)

# Sample of filled test cases
print("\n=== Sample filled test cases (Test Cases sheet) ===")
ws = wb["Test Cases"]
filled = 0
for r in range(5, ws.max_row + 1):
    tc = ws.cell(row=r, column=1).value
    if tc and str(tc).startswith("TC-"):
        status = ws.cell(row=r, column=7).value
        actual = ws.cell(row=r, column=8).value
        if status:
            filled += 1
            if filled <= 3 or filled in (8, 27):  # show first 3, plus FAIL + BLOCKED
                print(f"  {tc} [{status}] actual={actual!r}")
            if filled >= 35:
                break
print(f"Total rows with Status filled: {filled}")
