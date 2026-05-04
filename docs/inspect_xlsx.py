"""Inspect current state of ERP_Enterprise.xlsx — count test cases, sections, and read first rows."""
from openpyxl import load_workbook
import os

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "ERP_Enterprise.xlsx")

wb = load_workbook(XLSX, read_only=False)
print("Sheets:", wb.sheetnames)
ws = wb["Test Cases"]
print("Max row:", ws.max_row)
print("Max col:", ws.max_column)

# Header row is row 4
headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
print("Headers:", headers)

# Count actual test case rows (not section divider rows)
tc_rows = []
section_rows = []
for r in range(5, ws.max_row + 1):
    tc = ws.cell(row=r, column=1).value
    section = ws.cell(row=r, column=2).value
    flow = ws.cell(row=r, column=3).value
    if tc and str(tc).startswith("TC-"):
        tc_rows.append((r, tc, section, flow))
    elif tc and str(tc).startswith("§"):
        section_rows.append((r, tc))

print(f"Total test cases: {len(tc_rows)}")
print(f"Total section dividers: {len(section_rows)}")
print()
print("Sections found:")
for r, name in section_rows:
    print(f"  row {r}: {name}")
print()
print("First 10 test cases:")
for r, tc, section, flow in tc_rows[:10]:
    print(f"  {tc}: {section} / {flow}")
print()
print("Last 5 test cases:")
for r, tc, section, flow in tc_rows[-5:]:
    print(f"  {tc}: {section} / {flow}")
