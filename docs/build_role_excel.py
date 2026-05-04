"""Generate ONE clean single-sheet Excel for any role's test-flow MD.

Usage:
    python docs/build_role_excel.py <role>

<role> is one of: superadmin, reseller (case-insensitive).
Reads docs/ERP_<Role>.md and writes docs/ERP_<Role>.xlsx.
"""
from __future__ import annotations
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation


HERE = os.path.dirname(os.path.abspath(__file__))


# Per-role config: source filename + accent colour scheme
ROLES = {
    "superadmin": {
        "md_name": "ERP_SuperAdmin.md",
        "out_name": "ERP_SuperAdmin.xlsx",
        "title": "ERP — Super Admin · Test Cases",
        "title_fill": "4C1D95",   # violet-900
        "section_text": "4C1D95",
        "section_fill": "EDE9FE",  # violet-100
    },
    "reseller": {
        "md_name": "ERP_Reseller.md",
        "out_name": "ERP_Reseller.xlsx",
        "title": "ERP — Reseller · Test Cases",
        "title_fill": "B45309",   # amber-700
        "section_text": "B45309",
        "section_fill": "FEF3C7",  # amber-100
    },
}


def strip_inline(s: str) -> str:
    s = s.replace("\r", "")
    s = re.sub(r"`([^`]+)`", r"\1", s)
    s = re.sub(r"\*\*([^*]+)\*\*", r"\1", s)
    s = re.sub(r"\*([^*]+)\*", r"\1", s)
    s = re.sub(r"\[([^\]]+)\]\(([^)]+)\)", r"\1", s)
    s = s.replace("<br/>", "\n").replace("<br>", "\n")
    return s.strip()


def parse_table_row(line: str):
    p = line.strip()
    if p.startswith("|"):
        p = p[1:]
    if p.endswith("|"):
        p = p[:-1]
    return [strip_inline(c.strip()) for c in p.split("|")]


def extract_test_rows(path: str):
    with open(path, "r", encoding="utf-8") as f:
        lines = f.read().split("\n")

    section_idx = 0
    section_title = ""
    flow_title = ""
    last_para = ""
    last_h4 = ""
    in_code = False

    i = 0
    rows = []
    while i < len(lines):
        ln = lines[i]
        if ln.strip().startswith("```"):
            in_code = not in_code
            i += 1
            continue
        if in_code:
            i += 1
            continue

        m = re.match(r"^(#{1,6})\s+(.*)$", ln)
        if m:
            level = len(m.group(1))
            text = strip_inline(m.group(2))
            if level == 2:
                section_idx += 1
                section_title = text
                flow_title = ""
                last_para = ""
                last_h4 = ""
            elif level == 3:
                flow_title = text
                last_para = ""
                last_h4 = ""
            elif level == 4:
                last_h4 = text
            i += 1
            continue

        if ln.strip() and not ln.lstrip().startswith("|") and not ln.lstrip().startswith(("- ", "* ", ">")):
            buf = [ln.strip()]
            i += 1
            while i < len(lines) and lines[i].strip() and \
                    not lines[i].lstrip().startswith("|") and \
                    not re.match(r"^#{1,6}\s", lines[i]) and \
                    not lines[i].lstrip().startswith(("- ", "* ", ">")):
                buf.append(lines[i].strip())
                i += 1
            text = strip_inline(" ".join(buf))
            if len(text) < 200:
                last_para = text
            continue

        if ln.lstrip().startswith("|"):
            if i + 1 < len(lines) and re.match(r"^\s*\|?\s*[:\- ]+\|", lines[i + 1]):
                headers = parse_table_row(ln)
                i += 2
                table_rows = []
                while i < len(lines) and lines[i].lstrip().startswith("|"):
                    table_rows.append(parse_table_row(lines[i]))
                    i += 1

                flow_for_row = flow_title or section_title

                low = [h.lower() for h in headers]
                step_col = None
                expected_col = None
                for idx, h in enumerate(low):
                    if step_col is None and h in ("#", "step", "action", "field", "scenario", "persona"):
                        step_col = idx
                    if expected_col is None and ("expected" in h or "outcome" in h or "result" in h):
                        expected_col = idx
                if expected_col is None:
                    expected_col = len(headers) - 1
                if step_col is None:
                    step_col = 0

                for tr in table_rows:
                    if len(tr) < len(headers):
                        tr = tr + [""] * (len(headers) - len(tr))
                    step_label = tr[step_col] if step_col < len(tr) else ""
                    expected = tr[expected_col] if expected_col < len(tr) else ""
                    middle_parts = []
                    for j, h in enumerate(headers):
                        if j in (step_col, expected_col):
                            continue
                        cell = tr[j] if j < len(tr) else ""
                        if cell.strip():
                            middle_parts.append(f"{headers[j]}: {cell}")
                    action_text = "\n".join(middle_parts)
                    if len(headers) == 2:
                        action_text = step_label
                        step_label = ""
                    rows.append({
                        "section": section_title,
                        "flow": flow_for_row + (f" → {last_h4}" if last_h4 else ""),
                        "step": step_label,
                        "action": action_text,
                        "expected": expected,
                        "preconditions": last_para if "pre-condition" in last_para.lower() or "url" in last_para.lower()[:20] else "",
                    })
                continue

        i += 1

    return rows


def write_workbook(rows, out_path, cfg):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    title_font = Font(bold=True, color="FFFFFF", size=14)
    title_fill = PatternFill("solid", fgColor=cfg["title_fill"])
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F2937")
    section_font = Font(bold=True, color=cfg["section_text"], size=12)
    section_fill = PatternFill("solid", fgColor=cfg["section_fill"])
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    even_fill = PatternFill("solid", fgColor="F9FAFB")
    border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = cfg["title"]
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = (
        "Single-sheet view. One row per test step. "
        "Cross-role isolation failures = P0 SECURITY. "
        f"Source: docs/{os.path.basename(cfg['md_name'])}."
    )
    c.font = Font(italic=True, color="6B7280", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = [
        "TC #", "Section", "Flow", "Step", "Action / Test Data",
        "Expected Result", "Status", "Actual Result", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = "A5"

    cur_section = None
    excel_row = 5
    tc_num = 0
    for r in rows:
        if r["section"] != cur_section:
            cur_section = r["section"]
            ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=9)
            sc = ws.cell(row=excel_row, column=1, value=f"§ {cur_section}")
            sc.font = section_font
            sc.fill = section_fill
            sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            sc.border = border
            ws.row_dimensions[excel_row].height = 22
            excel_row += 1
        tc_num += 1
        values = [
            f"TC-{tc_num:04d}",
            r["section"],
            r["flow"],
            r["step"],
            r["action"],
            r["expected"],
            "",
            "",
            r["preconditions"] or "",
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col, value=v)
            cell.alignment = wrap
            cell.border = border
            cell.fill = even_fill if (excel_row % 2 == 0) else odd_fill
        ws.cell(row=excel_row, column=1).font = Font(bold=True, color="111827")
        ws.cell(row=excel_row, column=4).alignment = center
        excel_row += 1
    last_row = excel_row - 1

    widths = {"A": 11, "B": 32, "C": 38, "D": 8, "E": 50, "F": 60, "G": 13, "H": 35, "I": 30}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    pass_fill = PatternFill("solid", fgColor="D1FAE5")
    fail_fill = PatternFill("solid", fgColor="FECACA")
    blocked_fill = PatternFill("solid", fgColor="FEF3C7")
    skip_fill = PatternFill("solid", fgColor="E5E7EB")
    rng = f"G5:G{last_row}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=['EXACT(UPPER($G5),"PASS")'], fill=pass_fill))
    ws.conditional_formatting.add(rng, FormulaRule(formula=['EXACT(UPPER($G5),"FAIL")'], fill=fail_fill))
    ws.conditional_formatting.add(rng, FormulaRule(formula=['EXACT(UPPER($G5),"BLOCKED")'], fill=blocked_fill))
    ws.conditional_formatting.add(rng, FormulaRule(formula=['EXACT(UPPER($G5),"SKIP")'], fill=skip_fill))

    dv = DataValidation(type="list", formula1='"PASS,FAIL,BLOCKED,SKIP"', allow_blank=True)
    dv.add(rng)
    ws.add_data_validation(dv)

    ws.auto_filter.ref = f"A4:I{last_row}"

    fallback = out_path.replace(".xlsx", ".locked.xlsx")
    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        wb.save(fallback)
        return fallback


def run_for_role(role_key: str):
    if role_key not in ROLES:
        raise SystemExit(f"Unknown role '{role_key}'. Use one of: {', '.join(ROLES)}")
    cfg = ROLES[role_key]
    md_path = os.path.join(HERE, cfg["md_name"])
    out_path = os.path.join(HERE, cfg["out_name"])
    rows = extract_test_rows(md_path)
    saved = write_workbook(rows, out_path, cfg)
    print(f"[{role_key}] {len(rows)} rows -> {saved}")


def main():
    if len(sys.argv) < 2:
        # Run both by default
        for k in ROLES:
            run_for_role(k)
    else:
        run_for_role(sys.argv[1].lower())


if __name__ == "__main__":
    main()
